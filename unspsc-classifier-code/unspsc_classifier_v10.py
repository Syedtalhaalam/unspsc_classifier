#!/usr/bin/env python3
# -*- coding: utf-8 -*-
"""
UNSPSC Classification Pipeline - Production Ready v1.0
=========================================================

Exact implementation of user requirements for 90,000 materials classification.

WORKFLOW:
---------
Step 1: Read Excel with Material, Material_Description, PO_Text, Challenge_Level
Step 2: Create expert prompt for UNSPSC classification
Step 3: Send to 3 LLMs (ChatGPT, Gemini, Claude) - get top 10 each
Step 4: Find common codes, fill to 5 with ChatGPT (by confidence)
Step 5: Rerank 5 codes with all 3 LLMs
Step 6: Apply consensus (3/3, 2/3, or ChatGPT fallback)
Step 7: Output comprehensive Excel with all results

FEATURES:
---------
- Parallel API calls for speed
- JSONL caching to avoid reprocessing
- Error handling and retry logic
- Progress tracking for 90K materials
- Configurable rate limiting
- Production-grade logging
"""

from __future__ import annotations

import argparse
import hashlib
import json
import logging
import os
import re
import sys
import time
from concurrent.futures import ThreadPoolExecutor, as_completed
from dataclasses import dataclass, field
from datetime import datetime
from pathlib import Path
from typing import Any, Dict, List, Optional, Set, Tuple

import pandas as pd
from tqdm import tqdm

# API clients
try:
    import openai
    from openai import OpenAI
except ImportError:
    print("ERROR: openai package not installed. Run: pip install openai")
    sys.exit(1)

try:
    import anthropic
except ImportError:
    print("ERROR: anthropic package not installed. Run: pip install anthropic")
    sys.exit(1)

try:
    import google.generativeai as genai
    from google.generativeai.types import HarmCategory, HarmBlockThreshold
    from google.generativeai.types.generation_types import StopCandidateException
except ImportError:
    print("ERROR: google-generativeai package not installed. Run: pip install google-generativeai")
    sys.exit(1)


# =============================================================================
# Default Configuration Paths
# =============================================================================

# Configure your default paths here
DEFAULT_PATHS = {
    "input": r"D:\UNSPSC_Project\data\materials.xlsx",
    "unspsc": r"D:\UNSPSC_Project\data\unspsc_catalog.csv",
    "output": r"D:\UNSPSC_Project\output\results.xlsx",
    "cache": r"D:\UNSPSC_Project\cache\cache.jsonl",
}

# Optional: Model configurations
DEFAULT_MODELS = {
    "gpt": "gpt-4.1-2025-04-14",          # or "gpt-4.1-2025-04-14" for better quality
    "gemini": "gemini-2.0-flash-001",  # or "gemini-pro"
    "claude": "claude-haiku-4-5-20251001",  # or "claude-opus-4" for best quality
}
#gpt-5-2025-08-07 (reasoning Model)
#gpt-4.1-mini-2025-04-14
#gpt-4.1-2025-04-14
#gpt-5-2025-08-07


# Optional: Processing configurations
DEFAULT_CONFIG = {
    "parallel": True,              # Enable parallel API calls (faster)
    "cache_enabled": True,         # Enable caching (recommended for large batches)
    "verbose": False,              # Detailed logging
    "material_col": "Material",
    "description_col": "Material_Description",
    "po_col": "PO_Text",
    "challenge_col": "Challenge_Level",
    "item_detail_col": "Item_detail",
    "matl_group_col": "Matl Group",
    "plant_col": "Plant",
}


# =============================================================================
# Configuration & Logging
# =============================================================================

def setup_logging(verbose: bool = False) -> None:
    """Configure logging with appropriate level and format."""
    level = logging.DEBUG if verbose else logging.INFO
    logging.basicConfig(
        level=level,
        format="%(asctime)s | %(levelname)-8s | %(name)s | %(message)s",
        datefmt="%Y-%m-%d %H:%M:%S",
        handlers=[
            logging.FileHandler("unspsc_pipeline.log"),
            logging.StreamHandler(sys.stdout)
        ]
    )

logger = logging.getLogger(__name__)


# =============================================================================
# Data Classes
# =============================================================================

@dataclass
class Candidate:
    """Single UNSPSC candidate recommendation."""
    code: str
    commodity: str
    confidence: float
    rationale: str = ""
    provider: str = ""  # Track which LLM this came from


@dataclass
class ProviderResponse:
    """Response from a single LLM provider."""
    provider: str
    candidates: List[Candidate] = field(default_factory=list)
    error: Optional[str] = None
    raw_response: str = ""
    timestamp: str = field(default_factory=lambda: datetime.now().isoformat())


@dataclass
class ClassificationResult:
    """Complete classification result for a single material."""
    material_code: str
    material_description: str
    po_text: str
    challenge_level: str
    item_detail: str = ""  # NEW: Additional detail about the material
    matl_group: str = ""  # Material Group
    plant: str = ""  # Plant
    
    # Tier 1: Candidate generation (top 10 from each)
    gpt_candidates: List[Candidate] = field(default_factory=list)
    gemini_candidates: List[Candidate] = field(default_factory=list)
    claude_candidates: List[Candidate] = field(default_factory=list)
    
    # NEW V8: Top 4 commodities with codes from each LLM
    gpt_top4_with_codes: List[str] = field(default_factory=list)
    gemini_top4_with_codes: List[str] = field(default_factory=list)
    claude_top4_with_codes: List[str] = field(default_factory=list)
    
    # NEW V8: All commodities ranked by confidence
    all_commodities_ranked: List[str] = field(default_factory=list)
    
    # Aggregation
    common_codes: List[str] = field(default_factory=list)
    top_5_recommendations: List[str] = field(default_factory=list)
    shortlist_for_rerank: List[str] = field(default_factory=list)
    all_30_ranked: List[str] = field(default_factory=list)  # NEW: All 30 commodities ranked by confidence
    
    # Tier 2: Reranking results
    final_gpt_code: Optional[str] = None
    final_gemini_code: Optional[str] = None
    final_claude_code: Optional[str] = None
    
    # Final consensus
    final_recommendation: Optional[str] = None
    final_commodity: Optional[str] = None
    decision_rule: str = ""
    confidence: float = 0.0
    
    # UNSPSC hierarchy
    segment: Optional[str] = None
    family: Optional[str] = None
    class_name: Optional[str] = None
    
    # Metadata
    processing_time: float = 0.0
    from_cache: bool = False
    timestamp: str = field(default_factory=lambda: datetime.now().strftime("%Y-%m-%d"))


# =============================================================================
# UNSPSC Dictionary
# =============================================================================

class UNSPSCDict:
    """Fast lookup dictionary for UNSPSC codes and hierarchy."""
    
    def __init__(self, csv_path: Optional[str]):
        self.map: Dict[str, Dict[str, str]] = {}
        
        if not csv_path or not Path(csv_path).exists():
            logger.warning(f"UNSPSC file not found: {csv_path}. Will operate without validation.")
            return
            
        try:
            df = pd.read_csv(csv_path, dtype=str)
            logger.info(f"Loading UNSPSC codes from {csv_path}")
            
            # Detect column names (flexible)
            code_col = self._find_column(df, ['code', 'unspsc', 'unspsc_code'])
            title_col = self._find_column(df, ['title', 'name', 'description'])
            segment_col = self._find_column(df, ['segment'])
            family_col = self._find_column(df, ['family'])
            class_col = self._find_column(df, ['class'])
            
            if not code_col:
                code_col = df.columns[0]
            if not title_col:
                title_col = df.columns[1] if len(df.columns) > 1 else df.columns[0]
            
            logger.info(f"Using columns: code={code_col}, title={title_col}")
            
            for _, row in df.iterrows():
                code = self._normalize_code(row[code_col])
                title = self._clean_text(row[title_col])
                
                # Handle "code - title" format
                if title and re.match(r"^\d{8}\s*-\s*", title):
                    title = self._clean_text(title.split('-', 1)[-1])
                
                if code and title:
                    self.map[code] = {
                        'title': title,
                        'segment': self._clean_text(row[segment_col]) if segment_col else None,
                        'family': self._clean_text(row[family_col]) if family_col else None,
                        'class': self._clean_text(row[class_col]) if class_col else None,
                    }
            
            logger.info(f"Loaded {len(self.map):,} UNSPSC codes")
            
        except Exception as e:
            logger.error(f"Error loading UNSPSC file: {e}")
            
    def _find_column(self, df: pd.DataFrame, possible_names: List[str]) -> Optional[str]:
        """Find column by possible names (case-insensitive)."""
        for col in df.columns:
            col_lower = col.strip().lower().replace(" ", "_")
            if col_lower in possible_names:
                return col
        return None
    
    def _normalize_code(self, code: Any) -> Optional[str]:
        """Extract and normalize 8-digit UNSPSC code."""
        if code is None or pd.isna(code):
            return None
        s = str(code).strip().split()[0].split('-')[0]
        s = re.sub(r'\D', '', s)
        match = re.search(r'(\d{8})', s)
        return match.group(1) if match else None
    
    def _clean_text(self, text: Any) -> Optional[str]:
        """Clean and normalize text."""
        if text is None or pd.isna(text):
            return None
        return str(text).strip()
    
    def get_title(self, code: str) -> Optional[str]:
        """Get commodity title for code."""
        return self.map.get(code, {}).get('title')
    
    def get_hierarchy(self, code: str) -> Dict[str, Optional[str]]:
        """Get full hierarchy for code."""
        return self.map.get(code, {
            'title': None,
            'segment': None,
            'family': None,
            'class': None
        })
    
    def search_by_commodity_name(self, commodity_name: str, limit: int = 10) -> List[Tuple[str, str, float]]:
        """Search for UNSPSC codes by commodity name.
        
        Args:
            commodity_name: The commodity name to search for
            limit: Maximum number of results to return
            
        Returns:
            List of tuples (code, title, similarity_score) sorted by relevance
        """
        if not commodity_name or not self.map:
            return []
        
        # Normalize search term
        search_term = commodity_name.lower().strip()
        
        # Find matches
        matches = []
        for code, data in self.map.items():
            title = data['title'].lower()
            
            # Calculate similarity score
            if search_term == title:
                # Exact match
                score = 1.0
            elif search_term in title:
                # Substring match - score based on position and length
                score = 0.8 - (title.index(search_term) * 0.01)
            elif title in search_term:
                # Reverse substring match
                score = 0.7
            else:
                # Check word overlap
                search_words = set(search_term.split())
                title_words = set(title.split())
                common_words = search_words & title_words
                if common_words:
                    score = 0.5 * (len(common_words) / max(len(search_words), len(title_words)))
                else:
                    continue
            
            matches.append((code, data['title'], score))
        
        # Sort by score descending
        matches.sort(key=lambda x: x[2], reverse=True)
        
        return matches[:limit]
    
    def get_code_by_commodity_name(self, commodity_name: str) -> Optional[str]:
        """Get the best matching UNSPSC code for a commodity name.
        
        Args:
            commodity_name: The commodity name to search for
            
        Returns:
            UNSPSC code or None if no match found
        """
        matches = self.search_by_commodity_name(commodity_name, limit=1)
        return matches[0][0] if matches else None
    
    def exists(self, code: str) -> bool:
        """Check if code exists in dictionary."""
        return code in self.map


# =============================================================================
# JSONL Cache
# =============================================================================

class JSONLCache:
    """Simple JSONL cache for classification results."""
    
    def __init__(self, cache_file: Optional[str]):
        self.cache_file = cache_file
        self.cache: Dict[str, Dict] = {}
        
        if cache_file and Path(cache_file).exists():
            self._load()
    
    def _load(self):
        """Load cache from JSONL file."""
        try:
            with open(self.cache_file, 'r', encoding='utf-8') as f:
                for line in f:
                    if line.strip():
                        data = json.loads(line)
                        key = data.get('_cache_key')
                        if key:
                            self.cache[key] = data
            logger.info(f"Loaded {len(self.cache)} cached results from {self.cache_file}")
        except Exception as e:
            logger.error(f"Error loading cache: {e}")
    
    def get(self, key: str) -> Optional[Dict]:
        """Get cached result by key."""
        return self.cache.get(key)
    
    def _make_json_serializable(self, obj):
        """Convert objects to JSON-serializable format."""
        if isinstance(obj, list):
            return [self._make_json_serializable(item) for item in obj]
        elif isinstance(obj, dict):
            return {k: self._make_json_serializable(v) for k, v in obj.items()}
        elif hasattr(obj, '__dict__'):
            # Convert dataclass/object to dict
            return self._make_json_serializable(obj.__dict__)
        elif isinstance(obj, (str, int, float, bool, type(None))):
            return obj
        else:
            # Fallback: convert to string
            return str(obj)
    
    def set(self, key: str, data: Dict):
        """Save result to cache."""
        # Convert to JSON-serializable format
        serializable_data = self._make_json_serializable(data)
        serializable_data['_cache_key'] = key
        self.cache[key] = serializable_data
        
        if self.cache_file:
            try:
                with open(self.cache_file, 'a', encoding='utf-8') as f:
                    f.write(json.dumps(serializable_data, ensure_ascii=False) + '\n')
            except Exception as e:
                logger.error(f"Error writing to cache: {e}")
                logger.debug(f"Data type causing issue: {type(data)}")
    
    def make_key(self, material: str, po_text: str, item_detail: str = "") -> str:
        """Create cache key from inputs."""
        combined = f"{material}|{po_text}|{item_detail}".lower().strip()
        return hashlib.md5(combined.encode('utf-8')).hexdigest()


# =============================================================================
# Prompt Templates
# =============================================================================

class PromptTemplates:
    """Prompt templates for UNSPSC classification."""
    

#     TIER1_SYSTEM = """You are a UNSPSC classification engine for a large electrical utility company in power generation, transmission, and distribution.

# Your task is to classify materials using ONLY official UNSPSC codes (version 24.XX or later) by analyzing the material’s **primary function** and usage context.

# STRICT CLASSIFICATION RULES:

# 1. ✅ Return only official 8-digit UNSPSC codes from the latest UNSPSC version.
# 2. ❌ Never fabricate or invent a code. If no suitable match exists, return:
#    {"unspsc_code": "UNKNOWN", "commodity": "UNKNOWN", "confidence": 0.0, "rationale": "Explanation"}
# 3. ✅ Cross-check that the **commodity name matches** the code exactly from the UNSPSC dictionary.
# 4. ✅ The material’s **nature/type** is defined by the word before the first semicolon (`;`) in the `Material_Description`. This is your PRIMARY classification anchor.
#    - Example: For `CABLE; AL XLPE LT ABC 2X10 MM2`, classify the **CABLE**, not the insulation.
#    - Example: For `BOX; FUSE ICTP 400A LT DB`, classify the **BOX**, not the breaker.
# 5. ✅ Use `PO_Text` only if additional context or disambiguation is needed (voltage rating, function, configuration, etc.).
# 6. ❌ Do NOT classify accessories, packaging, or mounting hardware unless they are the **primary function**.
# 7. ✅ Always return the most specific 8-digit code available. Avoid 2/4/6-digit parent codes or any ending in ‘00’.
# 8. ✅ Prioritize electrical-utility context: transformers, LV/MV/HV switchgear, breakers, ABC, cables, CTs, relays, terminals.
# 9. ✅ Use item specifications to disambiguate when needed: current (A), voltage (V/kV), size (mm²), pressure (bar/psi), power (kVA/MVA).
# 10. ✅ If you’re not confident, choose the **closest valid code** and clearly lower the confidence.

# OUTPUT FORMAT (mandatory – do not change structure):

# {
#   "candidates": [
#     {
#       "unspsc_code": "XXXXXXXX",
#       "commodity": "Commodity Title",
#       "confidence": 0.92,
#       "rationale": "Short, clear justification: primary item + function + keyword + spec"
#     },
#     ...
#   ]
# }

# Return only top 10 most likely candidates. No explanations, markdown, or surrounding text.
# """

    TIER1_SYSTEM = """You are an expert UNSPSC classification engine for an electrical utility company in power generation, transmission, and distribution. Your goal is to classify each material based on its **primary function**, returning the most accurate **8-digit UNSPSC code** and commodity title.

CLASSIFICATION RULES (STRICT)

1. Return only **valid 8-digit UNSPSC codes and commodity titles** from the latest version (v24.x+).
2. Never fabricate or invent a code and commodity titles. If no suitable match exists.
3. Please ensure that the commodity name returned is correct. For example, the code 39121603 was given with the commodity name (Molded case circuit breakers) but the correct UNSPSC code for Molded case circuit breakers is 39121616 not 39121603.
4. Use `Item_Detail` field as additional context to understand the material's function, purpose, and usage. This field provides important details that can help disambiguate materials and improve classification accuracy.
5. Assign confidence levels as follows:
   - High (0.95+) if rule match is exact and historical match is unique
   - Medium (0.75–0.90) if description hints match, but multiple codes possible
   - Low (<0.70) if match is weak or fallback applied

6. Do NOT return codes ending in ‘00’ or partial segment/class codes.
7. Prioritize electrical-utility context: substations, HV/MV switchgear, transformers, ABC cables, relays, etc.

OUTPUT FORMAT (strictly JSON):

```json
{
  "candidates": [
    {
      "unspsc_code": "XXXXXXXX",
      "commodity": "Official UNSPSC commodity name",
      "confidence": 0.92,
      "rationale": "Keyword = X, function = Y, spec = Z (based on rules)"
    }
  ]
}
"""

# Return top 10 codes."""

    @staticmethod
    def tier1_user_prompt(material_description: str, po_text: str, item_detail: str = "") -> str:
        """Create Tier 1 user prompt."""
        po_display = po_text if po_text and str(po_text).strip().lower() != 'nan' else '(not provided)'
        item_detail_display = item_detail if item_detail and str(item_detail).strip().lower() != 'nan' else '(not provided)'
        
        return f"""Material_Description: {material_description}
PO_Text: {po_display}
Item_Detail: {item_detail_display}

Analyze and return the top 10 most relevant UNSPSC codes."""

    # Step 5: Tier 2 prompt for reranking (pick best from shortlist)
#     TIER2_SYSTEM = """You are an expert UNSPSC classifier. You will be given a shortlist of candidate UNSPSC codes.

# Your task: Pick the SINGLE BEST code from this shortlist that most accurately represents the material.

# Consider:
# 1. Primary function and intended use
# 2. Technical specifications and context
# 3. Most specific classification available

# CRITICAL: You must respond with ONLY valid JSON. No explanations, no markdown, no additional text.

# OUTPUT FORMAT (valid JSON only):
# {{
#   "final_code": "XXXXXXXX",
#   "confidence": 0.90,
#   "rationale": "Why this code is the best match"
# }}

# Response must be valid JSON that can be parsed directly."""

    TIER2_SYSTEM = """I am an expert UNSPSC classifier for a large electrical utility company in power generation, transmission, and distribution.
Note: 1. Before providing any UNSPSC code, please ensure it is a valid and correct code that exists within the official UNSPSC dictionary or hierarchy. 

Your task: Pick the SINGLE BEST code from this shortlist that most accurately represents the material.

Consider:
1. Primary function and intended use
2. Technical specifications and context
3. Most specific classification available

CRITICAL: You must respond with ONLY valid JSON. No explanations, no markdown, no additional text.

OUTPUT (valid JSON only):
{
  "final_code": "XXXXXXXX",
  "confidence": 0.92,
  "rationale": "Why best"
}"""

    @staticmethod
    def tier2_user_prompt(material_description: str, po_text: str, shortlist: List[str], unspsc_dict: UNSPSCDict, item_detail: str = "") -> str:
        """Create Tier 2 reranking prompt."""
        po_display = po_text if po_text and str(po_text).strip().lower() != 'nan' else '(not provided)'
        item_detail_display = item_detail if item_detail and str(item_detail).strip().lower() != 'nan' else '(not provided)'
        
        # Format shortlist with titles
        shortlist_formatted = []
        for i, code in enumerate(shortlist, 1):
            title = unspsc_dict.get_title(code) or "(unknown)"
            shortlist_formatted.append(f"{i}. {code} - {title}")
        
        shortlist_text = "\n".join(shortlist_formatted)
        
        return f"""Material_Description: {material_description}
PO_Text: {po_display}
Item_Detail: {item_detail_display}

SHORTLIST OF CANDIDATES:
{shortlist_text}

Pick the SINGLE BEST code from this shortlist. Return JSON only."""


# =============================================================================
# LLM Providers
# =============================================================================

class LLMProvider:
    """Base class for LLM providers."""
    
    def __init__(self, name: str, api_key: Optional[str]):
        self.name = name
        self.api_key = api_key
    
    def is_available(self) -> bool:
        """Check if provider is available."""
        return bool(self.api_key)
    
    def call_tier1(self, system: str, user: str) -> ProviderResponse:
        """Call Tier 1 (candidate generation)."""
        raise NotImplementedError
    
    def call_tier2(self, system: str, user: str) -> ProviderResponse:
        """Call Tier 2 (reranking)."""
        raise NotImplementedError
    
    def _parse_json_response(self, text: str) -> Optional[Dict]:
        """Parse JSON from LLM response with robust error handling."""
        if not text or not text.strip():
            logger.error(f"{self.name}: Empty response")
            return None
        
        original_text = text
        
        try:
            # Step 1: Remove markdown code blocks
            text = re.sub(r'```json\s*', '', text, flags=re.IGNORECASE)
            text = re.sub(r'```\s*', '', text)
            text = text.strip()
            
            # Step 2: Try direct parse
            try:
                return json.loads(text)
            except json.JSONDecodeError:
                pass
            
            # Step 3: Extract JSON from text (sometimes LLMs add explanations)
            # Look for JSON object between { and }
            json_match = re.search(r'\{.*\}', text, re.DOTALL)
            if json_match:
                try:
                    return json.loads(json_match.group())
                except json.JSONDecodeError:
                    pass
            
            # Step 4: Try to find JSON array
            json_match = re.search(r'\[.*\]', text, re.DOTALL)
            if json_match:
                try:
                    data = json.loads(json_match.group())
                    # If it's a list, wrap in dict
                    if isinstance(data, list):
                        return {'candidates': data}
                except json.JSONDecodeError:
                    pass
            
            # Step 5: Fix common JSON issues
            # Replace single quotes with double quotes
            text_fixed = text.replace("'", '"')
            try:
                return json.loads(text_fixed)
            except json.JSONDecodeError:
                pass
            
            # If all else fails, log and return None
            logger.error(f"{self.name}: JSON parse error - could not extract valid JSON")
            logger.debug(f"Original text (first 500 chars): {original_text[:500]}")
            logger.debug(f"Processed text (first 500 chars): {text[:500]}")
            return None
            
        except Exception as e:
            logger.error(f"{self.name}: Unexpected error parsing JSON: {e}")
            logger.debug(f"Raw text: {original_text[:500]}")
            return None


class ChatGPTProvider(LLMProvider):
    """OpenAI ChatGPT provider."""
    
    def __init__(self, api_key: Optional[str], model: str = "gpt-4.1-2025-04-14"):
        super().__init__("chatgpt", api_key)
        self.model = model
        if api_key:
            self.client = OpenAI(api_key=api_key)
    
    def call_tier1(self, system: str, user: str) -> ProviderResponse:
        """Get top 10 candidates."""
        if not self.is_available():
            return ProviderResponse(provider=self.name, error="API key not available")
        
        try:
            response = self.client.chat.completions.create(
                model=self.model,
                messages=[
                    {"role": "system", "content": system},
                    {"role": "user", "content": user}
                ],
                temperature=0.3,
                max_tokens=2000
            )
            
            raw_text = response.choices[0].message.content
            data = self._parse_json_response(raw_text)
            
            if not data or 'candidates' not in data:
                return ProviderResponse(
                    provider=self.name,
                    error="Invalid JSON response",
                    raw_response=raw_text
                )
            
            candidates = [
                Candidate(
                    code=c.get('unspsc_code', ''),
                    commodity=c.get('commodity', ''),
                    confidence=float(c.get('confidence', 0.5)),
                    rationale=c.get('rationale', ''),
                    provider='chatgpt'
                )
                for c in data['candidates'][:10]
            ]
            
            return ProviderResponse(
                provider=self.name,
                candidates=candidates,
                raw_response=raw_text
            )
            
        except Exception as e:
            logger.error(f"ChatGPT Tier1 error: {e}")
            return ProviderResponse(provider=self.name, error=str(e))
    
    def call_tier2(self, system: str, user: str) -> ProviderResponse:
        """Pick best from shortlist."""
        if not self.is_available():
            return ProviderResponse(provider=self.name, error="API key not available")
        
        try:
            response = self.client.chat.completions.create(
                model=self.model,
                messages=[
                    {"role": "system", "content": system},
                    {"role": "user", "content": user}
                ],
                temperature=0.2,
                max_tokens=500
            )
            
            raw_text = response.choices[0].message.content
            data = self._parse_json_response(raw_text)
            
            if not data or 'final_code' not in data:
                return ProviderResponse(
                    provider=self.name,
                    error="Invalid JSON response",
                    raw_response=raw_text
                )
            
            # Create single candidate for final pick
            candidate = Candidate(
                code=data.get('final_code', ''),
                commodity=data.get('commodity', ''),
                confidence=float(data.get('confidence', 0.5)),
                rationale=data.get('rationale', ''),
                provider='chatgpt'
            )
            
            return ProviderResponse(
                provider=self.name,
                candidates=[candidate],
                raw_response=raw_text
            )
            
        except Exception as e:
            logger.error(f"ChatGPT Tier2 error: {e}")
            return ProviderResponse(provider=self.name, error=str(e))


class GeminiProvider(LLMProvider):
    """Google Gemini provider."""
    
    def __init__(self, api_key: Optional[str], model: str = "gemini-2.0-flash-001"):
        super().__init__("gemini", api_key)
        self.model = model
        if api_key:
            genai.configure(api_key=api_key)
            self.client = genai.GenerativeModel(
                model_name=model,
                safety_settings={
                    HarmCategory.HARM_CATEGORY_HARASSMENT: HarmBlockThreshold.BLOCK_NONE,
                    HarmCategory.HARM_CATEGORY_HATE_SPEECH: HarmBlockThreshold.BLOCK_NONE,
                    HarmCategory.HARM_CATEGORY_SEXUALLY_EXPLICIT: HarmBlockThreshold.BLOCK_NONE,
                    HarmCategory.HARM_CATEGORY_DANGEROUS_CONTENT: HarmBlockThreshold.BLOCK_NONE,
                }
            )
    
    def call_tier1(self, system: str, user: str) -> ProviderResponse:
        """Get top 10 candidates with retry logic and JSON schema."""
        if not self.is_available():
            return ProviderResponse(provider=self.name, error="API key not available")
        
        # Define JSON schema for structured output
        response_schema = {
            "type": "object",
            "properties": {
                "candidates": {
                    "type": "array",
                    "items": {
                        "type": "object",
                        "properties": {
                            "unspsc_code": {"type": "string"},
                            "commodity": {"type": "string"},
                            "confidence": {"type": "number"},
                            "rationale": {"type": "string"}
                        },
                        "required": ["unspsc_code", "commodity", "confidence", "rationale"]
                    }
                }
            },
            "required": ["candidates"]
        }
        
        max_retries = 3  # Increased from 2
        for attempt in range(max_retries):
            try:
                # On retry, add even more explicit JSON instruction
                if attempt > 0:
                    prompt = f"{system}\n\n⚠️ CRITICAL: You MUST respond with ONLY valid JSON matching the exact schema. No explanations, no markdown, no other text.\n\n{user}\n\n⚠️ Remember: Output ONLY the JSON object with 'candidates' array. Nothing else."
                    logger.debug(f"Gemini Tier1 retry attempt {attempt + 1}/{max_retries}")
                else:
                    prompt = f"{system}\n\n{user}"
                
                response = self.client.generate_content(
                    prompt,
                    generation_config=genai.GenerationConfig(
                        temperature=0.3,
                        max_output_tokens=2000,
                        response_mime_type="application/json",
                        response_schema=response_schema  # Force schema compliance
                    )
                )
                
                raw_text = response.text
                logger.debug(f"Gemini Tier1 response length: {len(raw_text)} chars")
                
                data = self._parse_json_response(raw_text)
                
                if not data or 'candidates' not in data:
                    if attempt < max_retries - 1:
                        logger.warning(f"Gemini Tier1 invalid JSON (attempt {attempt + 1}), retrying...")
                        time.sleep(0.5)  # Brief pause before retry
                        continue
                    logger.error(f"Gemini Tier1 failed after {max_retries} attempts")
                    logger.debug(f"Last response: {raw_text[:500]}")
                    return ProviderResponse(
                        provider=self.name,
                        error=f"Invalid JSON response after {max_retries} retries",
                        raw_response=raw_text[:500]
                    )
                
                candidates = [
                    Candidate(
                        code=c.get('unspsc_code', ''),
                        commodity=c.get('commodity', ''),
                        confidence=float(c.get('confidence', 0.5)),
                        rationale=c.get('rationale', ''),
                        provider='gemini'
                    )
                    for c in data['candidates'][:10]
                ]
                
                logger.debug(f"Gemini Tier1 success: parsed {len(candidates)} candidates")
                return ProviderResponse(
                    provider=self.name,
                    candidates=candidates,
                    raw_response=raw_text
                )
                
            except StopCandidateException as e:
                logger.error(f"Gemini safety block: {e}")
                return ProviderResponse(provider=self.name, error=f"Safety block: {e}")
            except Exception as e:
                logger.error(f"Gemini Tier1 error (attempt {attempt + 1}): {e}")
                if attempt < max_retries - 1:
                    logger.debug(f"Will retry...")
                    time.sleep(0.5)
                    continue
                return ProviderResponse(provider=self.name, error=str(e))
    
    def call_tier2(self, system: str, user: str) -> ProviderResponse:
        """Pick best from shortlist with retry and JSON schema."""
        if not self.is_available():
            return ProviderResponse(provider=self.name, error="API key not available")
        
        # Define JSON schema for Tier 2
        response_schema = {
            "type": "object",
            "properties": {
                "final_code": {"type": "string"},
                "commodity": {"type": "string"},
                "confidence": {"type": "number"},
                "rationale": {"type": "string"}
            },
            "required": ["final_code", "confidence", "rationale"]
        }
        
        max_retries = 3
        for attempt in range(max_retries):
            try:
                if attempt > 0:
                    prompt = f"{system}\n\n⚠️ CRITICAL: Respond with ONLY valid JSON. No other text.\n\n{user}\n\n⚠️ Output ONLY JSON with 'final_code', 'confidence', 'rationale'."
                    logger.debug(f"Gemini Tier2 retry attempt {attempt + 1}/{max_retries}")
                else:
                    prompt = f"{system}\n\n{user}"
                
                response = self.client.generate_content(
                    prompt,
                    generation_config=genai.GenerationConfig(
                        temperature=0.2,
                        max_output_tokens=500,
                        response_mime_type="application/json",
                        response_schema=response_schema
                    )
                )
                
                raw_text = response.text
                data = self._parse_json_response(raw_text)
                
                if not data or 'final_code' not in data:
                    if attempt < max_retries - 1:
                        logger.warning(f"Gemini Tier2 invalid JSON (attempt {attempt + 1}), retrying...")
                        time.sleep(0.5)
                        continue
                    return ProviderResponse(
                        provider=self.name,
                        error=f"Invalid JSON response after {max_retries} retries",
                        raw_response=raw_text[:500]
                    )
                
                candidate = Candidate(
                    code=data.get('final_code', ''),
                    commodity=data.get('commodity', ''),
                    confidence=float(data.get('confidence', 0.5)),
                    rationale=data.get('rationale', ''),
                    provider='gemini'
                )
                
                logger.debug(f"Gemini Tier2 success: {candidate.code}")
                return ProviderResponse(
                    provider=self.name,
                    candidates=[candidate],
                    raw_response=raw_text
                )
                
            except StopCandidateException as e:
                logger.error(f"Gemini safety block: {e}")
                return ProviderResponse(provider=self.name, error=f"Safety block: {e}")
            except Exception as e:
                logger.error(f"Gemini Tier2 error (attempt {attempt + 1}): {e}")
                if attempt < max_retries - 1:
                    logger.debug(f"Will retry...")
                    time.sleep(0.5)
                    continue
                return ProviderResponse(provider=self.name, error=str(e))


class ClaudeProvider(LLMProvider):
    """Anthropic Claude provider."""
    
    def __init__(self, api_key: Optional[str], model: str = "claude-haiku-4-5-20251001"):
        super().__init__("claude", api_key)
        self.model = model
        if api_key:
            self.client = anthropic.Anthropic(api_key=api_key)
    
    def call_tier1(self, system: str, user: str) -> ProviderResponse:
        """Get top 10 candidates."""
        if not self.is_available():
            return ProviderResponse(provider=self.name, error="API key not available")
        
        try:
            response = self.client.messages.create(
                model=self.model,
                max_tokens=2000,
                temperature=0.3,
                system=system,
                messages=[{"role": "user", "content": user}]
            )
            
            raw_text = response.content[0].text
            data = self._parse_json_response(raw_text)
            
            if not data or 'candidates' not in data:
                return ProviderResponse(
                    provider=self.name,
                    error="Invalid JSON response",
                    raw_response=raw_text
                )
            
            candidates = [
                Candidate(
                    code=c.get('unspsc_code', ''),
                    commodity=c.get('commodity', ''),
                    confidence=float(c.get('confidence', 0.5)),
                    rationale=c.get('rationale', ''),
                    provider='claude'  # Fixed from 'chatgpt'
                )
                for c in data['candidates'][:10]
            ]
            
            return ProviderResponse(
                provider=self.name,
                candidates=candidates,
                raw_response=raw_text
            )
            
        except Exception as e:
            logger.error(f"Claude Tier1 error: {e}")
            return ProviderResponse(provider=self.name, error=str(e))
    
    def call_tier2(self, system: str, user: str) -> ProviderResponse:
        """Pick best from shortlist."""
        if not self.is_available():
            return ProviderResponse(provider=self.name, error="API key not available")
        
        try:
            response = self.client.messages.create(
                model=self.model,
                max_tokens=500,
                temperature=0.2,
                system=system,
                messages=[{"role": "user", "content": user}]
            )
            
            raw_text = response.content[0].text
            data = self._parse_json_response(raw_text)
            
            if not data or 'final_code' not in data:
                return ProviderResponse(
                    provider=self.name,
                    error="Invalid JSON response",
                    raw_response=raw_text
                )
            
            candidate = Candidate(
                code=data.get('final_code', ''),
                commodity=data.get('commodity', ''),
                confidence=float(data.get('confidence', 0.5)),
                rationale=data.get('rationale', ''),
                provider='claude'
            )
            
            return ProviderResponse(
                provider=self.name,
                candidates=[candidate],
                raw_response=raw_text
            )
            
        except Exception as e:
            logger.error(f"Claude Tier2 error: {e}")
            return ProviderResponse(provider=self.name, error=str(e))


# =============================================================================
# Classification Pipeline
# =============================================================================

class UNSPSCPipeline:
    """Main UNSPSC classification pipeline."""
    
    def __init__(
        self,
        unspsc_dict: UNSPSCDict,
        cache: JSONLCache,
        chatgpt: ChatGPTProvider,
        gemini: GeminiProvider,
        claude: ClaudeProvider,
        parallel: bool = True
    ):
        self.unspsc = unspsc_dict
        self.cache = cache
        self.providers = {
            'chatgpt': chatgpt,
            'gemini': gemini,
            'claude': claude
        }
        self.parallel = parallel
    
    def _normalize_code(self, code: Any) -> Optional[str]:
        """Normalize UNSPSC code to 8 digits."""
        if not code:
            return None
        s = str(code).strip().split()[0].split('-')[0]
        s = re.sub(r'\D', '', s)
        match = re.search(r'(\d{8})', s)
        return match.group(1) if match else None
    
    def _extract_keyword(self, material_description: str) -> Optional[str]:
        """Extract the first word/phrase before semicolon from Material_Description.
        
        Returns:
            The keyword before semicolon, or None if no semicolon found
        """
        if not material_description or pd.isna(material_description):
            return None
        
        material_description = str(material_description).strip()
        
        # Check if there's a semicolon
        if ';' not in material_description:
            return None
        
        # Extract text before first semicolon
        keyword = material_description.split(';')[0].strip()
        
        return keyword if keyword else None
    
    def _generate_item_detail(self, keyword: str, material_description: str, po_text: str) -> str:
        """Generate shortest definition for the keyword using ChatGPT with context.
        
        Args:
            keyword: The extracted keyword to define
            material_description: Full material description for context
            po_text: Purchase order text for additional context
            
        Returns:
            Generated definition string, or empty string if generation fails
        """
        # Check if ChatGPT is available
        if 'chatgpt' not in self.providers or not self.providers['chatgpt'].is_available():
            logger.warning(f"ChatGPT not available for Item_detail generation")
            return ""
        
        # Clean up po_text
        po_display = po_text if po_text and str(po_text).strip().lower() != 'nan' else '(not provided)'
        
        # Create prompt for definition generation
        system_prompt = """You are an expert in electrical utility equipment and materials. 
Your task is to provide the SHORTEST, most concise definition of a keyword extracted from a material description.

Rules:
1. Use the provided context (Material_Description and PO_Text) to understand the specific usage
2. Keep the definition to 1-2 sentences maximum
3. Focus on the function and purpose in electrical utility context
4. Be specific to the type of item indicated by the context
5. Return ONLY the definition text, no additional explanations

Example:
Keyword: "BASE"
Context: "BASE; FUSE HRC LT 630 A (NH3)"
Definition: "Fuse base/holder that supports the fuse link within a panel or carrier."
"""
        
        user_prompt = f"""Keyword: "{keyword}"
Material_Description: {material_description}
PO_Text: {po_display}

Provide the shortest definition of this keyword based on the context above."""
        
        try:
            # Call ChatGPT with low temperature for consistent, focused output
            response = self.providers['chatgpt'].client.chat.completions.create(
                model=self.providers['chatgpt'].model,
                messages=[
                    {"role": "system", "content": system_prompt},
                    {"role": "user", "content": user_prompt}
                ],
                temperature=0.3,
                max_tokens=150  # Keep it short
            )
            
            definition = response.choices[0].message.content.strip()
            
            # Clean up any quotes or markdown
            definition = definition.strip('"\'`')
            
            logger.info(f"Generated Item_detail for keyword '{keyword}': {definition[:50]}...")
            return definition
            
        except Exception as e:
            logger.error(f"Failed to generate Item_detail for keyword '{keyword}': {e}")
            return ""
    
    def _extract_top4_commodities_with_codes(self, candidates: List[Candidate]) -> List[str]:
        """Extract top 4 commodity names and find their UNSPSC codes from catalog.
        
        Args:
            candidates: List of Candidate objects with commodity names and confidence
            
        Returns:
            List of strings in format: "code - commodity (confidence: X.XX)"
            Sorted by confidence descending
        """
        if not candidates:
            return []
        
        # Sort candidates by confidence descending
        sorted_candidates = sorted(candidates, key=lambda x: x.confidence, reverse=True)
        
        # Take top 4
        top4 = sorted_candidates[:4]
        
        results = []
        for candidate in top4:
            commodity_name = candidate.commodity
            confidence = candidate.confidence
            
            # Try to find the code in UNSPSC catalog
            # First, check if candidate already has a valid code
            code = self._normalize_code(candidate.code)
            if code and self.unspsc.exists(code):
                # Use the existing code
                title = self.unspsc.get_title(code) or commodity_name
                results.append(f"{code} - {title} (confidence: {confidence:.2f})")
            else:
                # Search by commodity name in catalog
                matches = self.unspsc.search_by_commodity_name(commodity_name, limit=1)
                if matches:
                    matched_code, matched_title, similarity = matches[0]
                    results.append(f"{matched_code} - {matched_title} (confidence: {confidence:.2f})")
                else:
                    # No match found in catalog, still include with original commodity name
                    results.append(f"NO_CODE - {commodity_name} (confidence: {confidence:.2f})")
        
        return results
    
    def _rank_all_commodities_by_confidence(
        self, 
        gpt_candidates: List[Candidate],
        gemini_candidates: List[Candidate],
        claude_candidates: List[Candidate],
        gpt_top4_with_codes: List[str],
        gemini_top4_with_codes: List[str],
        claude_top4_with_codes: List[str],
    ) -> List[str]:
        """Rank ALL commodities from all LLMs by confidence level.

        Includes:
        - GPT_Commodity, Gemini_Commodity, Claude_Commodity (raw candidates)
        - GPT_Top4_With_Codes, Gemini_Top4_With_Codes, Claude_Top4_With_Codes

        Returns:
            List of unique strings: "code - title (confidence: X.XX)"
            Sorted by confidence descending.
        """
        # Use a dict to track the HIGHEST confidence for each unique code
        code_confidence_map: Dict[str, Tuple[str, float]] = {}  # code -> (title, max_confidence)

        # ---- from raw candidates ----
        for candidate in gpt_candidates:
            code = self._normalize_code(candidate.code)
            if code and self.unspsc.exists(code):
                title = self.unspsc.get_title(code) or candidate.commodity
                if code not in code_confidence_map or candidate.confidence > code_confidence_map[code][1]:
                    code_confidence_map[code] = (title, candidate.confidence)

        for candidate in gemini_candidates:
            code = self._normalize_code(candidate.code)
            if code and self.unspsc.exists(code):
                title = self.unspsc.get_title(code) or candidate.commodity
                if code not in code_confidence_map or candidate.confidence > code_confidence_map[code][1]:
                    code_confidence_map[code] = (title, candidate.confidence)

        for candidate in claude_candidates:
            code = self._normalize_code(candidate.code)
            if code and self.unspsc.exists(code):
                title = self.unspsc.get_title(code) or candidate.commodity
                if code not in code_confidence_map or candidate.confidence > code_confidence_map[code][1]:
                    code_confidence_map[code] = (title, candidate.confidence)

        # ---- also fold in Top4_With_Codes lists ----
        def _update_from_top4(entries: List[str]) -> None:
            for entry in entries:
                if not entry or ' - ' not in entry:
                    continue

                # Expected format: "CODE - title (confidence: X.XX)"
                code_part, rest = entry.split(' - ', 1)
                code = self._normalize_code(code_part.strip())
                if not code or code == 'NO_CODE':
                    continue
                if not self.unspsc.exists(code):
                    continue

                # Try to extract confidence from the string
                m = re.search(r"confidence:\s*([0-9]*\.?[0-9]+)", entry)
                try:
                    confidence = float(m.group(1)) if m else 0.0
                except ValueError:
                    confidence = 0.0

                # Prefer catalog title; fall back to what's in the string
                title = self.unspsc.get_title(code)
                if not title:
                    title = rest.split('(confidence', 1)[0].strip()

                if code not in code_confidence_map or confidence > code_confidence_map[code][1]:
                    code_confidence_map[code] = (title, confidence)

        _update_from_top4(gpt_top4_with_codes)
        _update_from_top4(gemini_top4_with_codes)
        _update_from_top4(claude_top4_with_codes)

        # Sort by confidence descending
        sorted_codes = sorted(code_confidence_map.items(), key=lambda x: x[1][1], reverse=True)

        # Format output - UNIQUE codes with NO provider tag
        results: List[str] = []
        for code, (title, confidence) in sorted_codes:
            results.append(f"{code} - {title} (confidence: {confidence:.2f})")

        logger.debug(f"All_Commodities_Ranked: {len(results)} unique codes")
        return results

    def _call_providers_tier1(self, system: str, user: str) -> Dict[str, ProviderResponse]:
        """Call all providers for Tier 1 (candidate generation)."""
        results = {}
        
        if self.parallel:
            with ThreadPoolExecutor(max_workers=3) as executor:
                futures = {
                    executor.submit(provider.call_tier1, system, user): name
                    for name, provider in self.providers.items()
                    if provider.is_available()
                }
                
                for future in as_completed(futures):
                    provider_name = futures[future]
                    try:
                        results[provider_name] = future.result(timeout=60)
                    except Exception as e:
                        logger.error(f"{provider_name} Tier1 failed: {e}")
                        results[provider_name] = ProviderResponse(provider=provider_name, error=str(e))
        else:
            for name, provider in self.providers.items():
                if provider.is_available():
                    results[name] = provider.call_tier1(system, user)
        
        return results
    
    def _call_providers_tier2(self, system: str, user: str) -> Dict[str, ProviderResponse]:
        """Call all providers for Tier 2 (reranking)."""
        results = {}
        
        if self.parallel:
            with ThreadPoolExecutor(max_workers=3) as executor:
                futures = {
                    executor.submit(provider.call_tier2, system, user): name
                    for name, provider in self.providers.items()
                    if provider.is_available()
                }
                
                for future in as_completed(futures):
                    provider_name = futures[future]
                    try:
                        results[provider_name] = future.result(timeout=60)
                    except Exception as e:
                        logger.error(f"{provider_name} Tier2 failed: {e}")
                        results[provider_name] = ProviderResponse(provider=provider_name, error=str(e))
        else:
            for name, provider in self.providers.items():
                if provider.is_available():
                    results[name] = provider.call_tier2(system, user)
        
        return results
    
    def _find_common_codes(
        self,
        gpt_codes: Set[str],
        gemini_codes: Set[str],
        claude_codes: Set[str]
    ) -> List[str]:
        """Step 4: Find codes common across all 3 LLMs."""
        common = gpt_codes & gemini_codes & claude_codes
        return list(common)
    
    def _build_shortlist(
        self,
        common_codes: List[str],
        gpt_candidates: List[Candidate]
    ) -> List[str]:
        """Step 4: Build 5-item shortlist per user requirements.
        
        Scenarios:
        - 5 common: use 5 common
        - 4 common: use 4 common + 1 ChatGPT (highest confidence, not in common)
        - 3 common: use 3 common + 2 ChatGPT (highest confidence, not in common)
        - 2 common: use 2 common + 3 ChatGPT (highest confidence, not in common)
        - 1 common: use 1 common + 4 ChatGPT (highest confidence, not in common)
        - 0 common: use 5 ChatGPT (highest confidence)
        """
        shortlist = list(common_codes)
        common_set = set(common_codes)
        
        # Calculate how many we need from ChatGPT
        num_common = len(common_codes)
        num_needed = 5 - min(num_common, 5)
        
        if num_needed > 0:
            # Get ChatGPT candidates not in common, sorted by confidence
            gpt_not_common = [
                c for c in gpt_candidates
                if self._normalize_code(c.code) and 
                   self._normalize_code(c.code) not in common_set
            ]
            gpt_not_common.sort(key=lambda x: x.confidence, reverse=True)
            
            # Add top N from ChatGPT
            for candidate in gpt_not_common[:num_needed]:
                code = self._normalize_code(candidate.code)
                if code and code not in shortlist:
                    shortlist.append(code)
                if len(shortlist) >= 5:
                    break
        
        return shortlist[:5]
    
    def _build_top_5_recommendations(
        self,
        all_candidates: List[Candidate],
        gpt_top4_with_codes: List[str],
        gemini_top4_with_codes: List[str],
        claude_top4_with_codes: List[str],
    ) -> List[str]:
        """Build Top 5 Recommendations from ALL sources.

        Includes:
        - GPT_Commodity (all candidates)
        - Gemini_Commodity (all candidates)
        - Claude_Commodity (all candidates)
        - GPT_Top4_With_Codes (top 4)
        - Gemini_Top4_With_Codes (top 4)
        - Claude_Top4_With_Codes (top 4)

        Returns:
            Unique UNSPSC codes, sorted by confidence (highest first), limited to top 5.
        """
        # Collect all unique (code, confidence) pairs
        code_confidence_map: Dict[str, float] = {}  # code -> max confidence

        # ---- from raw candidates ----
        for candidate in all_candidates:
            code = self._normalize_code(candidate.code)
            if code and self.unspsc.exists(code):
                if code not in code_confidence_map or candidate.confidence > code_confidence_map[code]:
                    code_confidence_map[code] = candidate.confidence

        # ---- also fold in Top4_With_Codes lists ----
        def _update_from_top4(entries: List[str]) -> None:
            for entry in entries:
                if not entry or ' - ' not in entry:
                    continue

                # Expected format: "CODE - title (confidence: X.XX)"
                code_part, _ = entry.split(' - ', 1)
                code = self._normalize_code(code_part.strip())
                if not code or code == 'NO_CODE':
                    continue
                if not self.unspsc.exists(code):
                    continue

                # Extract confidence if present
                m = re.search(r"confidence:\s*([0-9]*\.?[0-9]+)", entry)
                try:
                    confidence = float(m.group(1)) if m else 0.0
                except ValueError:
                    confidence = 0.0

                if code not in code_confidence_map or confidence > code_confidence_map[code]:
                    code_confidence_map[code] = confidence

        _update_from_top4(gpt_top4_with_codes)
        _update_from_top4(gemini_top4_with_codes)
        _update_from_top4(claude_top4_with_codes)

        # Sort by confidence descending
        sorted_codes = sorted(code_confidence_map.items(), key=lambda x: x[1], reverse=True)

        # Take top 5 unique codes
        top_5 = [code for code, _ in sorted_codes[:5]]

        logger.debug(f"  Top 5 Recommendations from {len(code_confidence_map)} unique codes: {top_5}")
        return top_5


    def _rank_all_30_by_confidence(
        self,
        all_candidates: List[Candidate],
        top_5_codes: List[str] = None
    ) -> List[str]:
        """Rank all 30 candidates (10 from each LLM) by confidence.
        
        CHANGES:
        1. Validates codes against UNSPSC catalog (removes invalid codes)
        2. Excludes codes already in Top_5_Recommendations
        
        Returns list of unique codes with titles, sorted by confidence descending.
        Includes provider info and confidence in format:
        "code - title (confidence: 0.95, provider: chatgpt)"
        """
        # Convert top_5_codes to set for exclusion (if provided)
        exclude_codes = set(top_5_codes) if top_5_codes else set()
        
        # Create a dict to track best confidence for each code
        code_best: Dict[str, Tuple[float, str, str]] = {}  # code -> (confidence, provider, title)
        
        for candidate in all_candidates:
            code = self._normalize_code(candidate.code)
            if not code:
                continue
            
            # CHANGE 1: Validate code exists in UNSPSC catalog
            if not self.unspsc.exists(code):
                logger.debug(f"  Skipping invalid code: {code} (not in UNSPSC catalog)")
                continue
            
            # CHANGE 2: Skip if code is in Top 5 Recommendations
            if code in exclude_codes:
                logger.debug(f"  Skipping code {code} (already in Top 5)")
                continue
            
            title = self.unspsc.get_title(code) or candidate.commodity
            confidence = candidate.confidence
            provider = candidate.provider or 'unknown'
            
            # Keep the highest confidence for each code
            if code not in code_best or confidence > code_best[code][0]:
                code_best[code] = (confidence, provider, title)
        
        # Sort by confidence descending
        sorted_codes = sorted(
            code_best.items(),
            key=lambda x: x[1][0],  # Sort by confidence
            reverse=True
        )
        
        # Format as strings
        formatted = []
        for code, (confidence, provider, title) in sorted_codes:
            formatted.append(f"{code} - {title} (conf: {confidence:.2f}, {provider})")
        
        return formatted
    
    def _apply_consensus(
        self,
        gpt_code: Optional[str],
        gemini_code: Optional[str],
        claude_code: Optional[str]
    ) -> Tuple[Optional[str], str]:
        """Step 5: Apply consensus logic per user requirements.
        
        Scenarios:
        1. If all 3/3 agree → use that code
        2. If 2/3 agree → use that code (even if one LLM returned None)
        3. If all different → use ChatGPT (if available), else Gemini, else Claude
        
        IMPORTANT: If any LLM returns None, consensus should still work with remaining LLMs.
        """
        codes = [c for c in [gpt_code, gemini_code, claude_code] if c]
        
        if not codes:
            return None, "no_valid_codes"
        
        # Count votes among non-None codes
        from collections import Counter
        vote_counts = Counter(codes)
        most_common = vote_counts.most_common(1)[0]
        winner_code, winner_votes = most_common
        
        # 3/3 consensus (all three returned same code)
        if winner_votes == 3:
            return winner_code, "3_of_3_consensus"
        
        # 2/3 consensus (two LLMs agreed, including cases where one returned None)
        if winner_votes == 2:
            return winner_code, "2_of_3_consensus"
        
        # All different or only 1-2 LLMs responded with different codes
        # Priority: ChatGPT > Gemini > Claude
        if gpt_code:
            return gpt_code, "chatgpt_fallback"
        elif gemini_code:
            return gemini_code, "gemini_fallback"
        elif claude_code:
            return claude_code, "claude_fallback"
        
        # Fallback (should never reach here given earlier checks)
        return winner_code, "fallback_most_common"
    
    def _reconstruct_from_cache(self, cached: Dict) -> ClassificationResult:
        """Reconstruct ClassificationResult from cached dictionary."""
        # Helper to convert dict back to Candidate
        def dict_to_candidate(d: Dict) -> Candidate:
            if isinstance(d, dict):
                return Candidate(
                    code=d.get('code', ''),
                    commodity=d.get('commodity', ''),
                    confidence=d.get('confidence', 0.0),
                    rationale=d.get('rationale', ''),
                    provider=d.get('provider', '')
                )
            return d
        
        # Reconstruct candidate lists
        gpt_candidates = [dict_to_candidate(c) for c in cached.get('gpt_candidates', [])]
        gemini_candidates = [dict_to_candidate(c) for c in cached.get('gemini_candidates', [])]
        claude_candidates = [dict_to_candidate(c) for c in cached.get('claude_candidates', [])]
        
        # Create result object
        result = ClassificationResult(
            material_code=cached.get('material_code', ''),
            material_description=cached.get('material_description', ''),
            po_text=cached.get('po_text', ''),
            challenge_level=cached.get('challenge_level', 'Medium'),
            gpt_candidates=gpt_candidates,
            gemini_candidates=gemini_candidates,
            claude_candidates=claude_candidates,
            common_codes=cached.get('common_codes', []),
            top_5_recommendations=cached.get('top_5_recommendations', []),
            shortlist_for_rerank=cached.get('shortlist_for_rerank', []),
            all_30_ranked=cached.get('all_30_ranked', []),  # NEW
            final_gpt_code=cached.get('final_gpt_code'),
            final_gemini_code=cached.get('final_gemini_code'),
            final_claude_code=cached.get('final_claude_code'),
            final_recommendation=cached.get('final_recommendation'),
            final_commodity=cached.get('final_commodity'),
            decision_rule=cached.get('decision_rule', ''),
            confidence=cached.get('confidence', 0.0),
            segment=cached.get('segment'),
            family=cached.get('family'),
            class_name=cached.get('class_name'),
            processing_time=cached.get('processing_time', 0.0),
            from_cache=True,
            timestamp=cached.get('timestamp', '')
        )
        
        return result
    
    def classify_material(
        self,
        material_code: str,
        material_description: str,
        po_text: str,
        challenge_level: str = "Medium",
        item_detail: str = "",
        matl_group: str = "",
        plant: str = ""
    ) -> ClassificationResult:
        """Classify a single material through complete pipeline."""
        
        start_time = time.time()
        
        # AUTO-GENERATE Item_detail if empty (before cache check)
        if not item_detail or str(item_detail).strip().lower() in ['', 'nan', 'none']:
            logger.debug(f"{material_code}: Item_detail is empty, attempting auto-generation")
            
            # Extract keyword from Material_Description
            keyword = self._extract_keyword(material_description)
            
            if keyword:
                logger.info(f"{material_code}: Extracted keyword '{keyword}' from Material_Description")
                
                # Generate definition using ChatGPT
                generated_definition = self._generate_item_detail(keyword, material_description, po_text)
                
                if generated_definition:
                    item_detail = generated_definition
                    logger.info(f"{material_code}: Auto-generated Item_detail: {generated_definition[:80]}...")
                else:
                    logger.warning(f"{material_code}: Failed to generate Item_detail for keyword '{keyword}'")
            else:
                logger.debug(f"{material_code}: No semicolon found in Material_Description, proceeding without Item_detail")
        
        # Check cache (after item_detail is finalized)
        cache_key = self.cache.make_key(material_description, po_text, item_detail)
        cached = self.cache.get(cache_key)
        
        if cached:
            logger.debug(f"Cache hit for {material_code}")
            result = self._reconstruct_from_cache(cached)
            result.from_cache = True
            return result
        
        # Initialize result
        result = ClassificationResult(
            material_code=material_code,
            material_description=material_description,
            po_text=po_text,
            challenge_level=challenge_level,
            item_detail=item_detail,
            matl_group=matl_group,
            plant=plant
        )
        
        try:
            # TIER 1: Candidate Generation (get top 10 from each LLM)
            logger.debug(f"Tier 1: Generating candidates for {material_code}")
            
            tier1_system = PromptTemplates.TIER1_SYSTEM
            tier1_user = PromptTemplates.tier1_user_prompt(material_description, po_text, item_detail)
            
            tier1_responses = self._call_providers_tier1(tier1_system, tier1_user)
            
            # Extract candidates
            result.gpt_candidates = tier1_responses.get('chatgpt', ProviderResponse(provider='chatgpt')).candidates
            result.gemini_candidates = tier1_responses.get('gemini', ProviderResponse(provider='gemini')).candidates
            result.claude_candidates = tier1_responses.get('claude', ProviderResponse(provider='claude')).candidates
            
            # Log all codes before validation
            logger.debug(f"Raw candidates - GPT: {len(result.gpt_candidates)}, Gemini: {len(result.gemini_candidates)}, Claude: {len(result.claude_candidates)}")
            
            # NEW V8: Extract top 4 commodities with their codes from each LLM
            logger.debug(f"Extracting top 4 commodities with codes for each LLM")
            result.gpt_top4_with_codes = self._extract_top4_commodities_with_codes(result.gpt_candidates)
            result.gemini_top4_with_codes = self._extract_top4_commodities_with_codes(result.gemini_candidates)
            result.claude_top4_with_codes = self._extract_top4_commodities_with_codes(result.claude_candidates)
            
            # NEW V8: Rank all commodities from all LLMs by confidence
            logger.debug(f"Ranking all commodities by confidence")
            result.all_commodities_ranked = self._rank_all_commodities_by_confidence(
                result.gpt_candidates,
                result.gemini_candidates,
                result.claude_candidates,
                result.gpt_top4_with_codes,
                result.gemini_top4_with_codes,
                result.claude_top4_with_codes,
            )
            
            # Normalize codes and validate against UNSPSC catalog
            gpt_codes_raw = {self._normalize_code(c.code) for c in result.gpt_candidates if self._normalize_code(c.code)}
            gemini_codes_raw = {self._normalize_code(c.code) for c in result.gemini_candidates if self._normalize_code(c.code)}
            claude_codes_raw = {self._normalize_code(c.code) for c in result.claude_candidates if self._normalize_code(c.code)}
            
            logger.debug(f"Normalized codes - GPT: {gpt_codes_raw}, Gemini: {gemini_codes_raw}, Claude: {claude_codes_raw}")
            
            # Filter out invalid codes
            gpt_codes = {code for code in gpt_codes_raw if self.unspsc.exists(code)}
            gemini_codes = {code for code in gemini_codes_raw if self.unspsc.exists(code)}
            claude_codes = {code for code in claude_codes_raw if self.unspsc.exists(code)}
            
            # Log invalid codes
            gpt_invalid = gpt_codes_raw - gpt_codes
            gemini_invalid = gemini_codes_raw - gemini_codes
            claude_invalid = claude_codes_raw - claude_codes
            
            if gpt_invalid:
                logger.warning(f"GPT invalid codes removed: {gpt_invalid}")
            if gemini_invalid:
                logger.warning(f"Gemini invalid codes removed: {gemini_invalid}")
            if claude_invalid:
                logger.warning(f"Claude invalid codes removed: {claude_invalid}")
            
            logger.info(f"Valid codes after filtering - GPT: {len(gpt_codes)}, Gemini: {len(gemini_codes)}, Claude: {len(claude_codes)}")
            
            # NEW V8: Extract codes from top 4 of each LLM (these will be prioritized)
            gpt_top4_codes = set()
            gemini_top4_codes = set()
            claude_top4_codes = set()
            
            for entry in result.gpt_top4_with_codes:
                # Extract code from format "CODE - title (confidence: X.XX)"
                if entry and ' - ' in entry:
                    code = entry.split(' - ')[0].strip()
                    if code != 'NO_CODE' and self.unspsc.exists(code):
                        gpt_top4_codes.add(code)
            
            for entry in result.gemini_top4_with_codes:
                if entry and ' - ' in entry:
                    code = entry.split(' - ')[0].strip()
                    if code != 'NO_CODE' and self.unspsc.exists(code):
                        gemini_top4_codes.add(code)
            
            for entry in result.claude_top4_with_codes:
                if entry and ' - ' in entry:
                    code = entry.split(' - ')[0].strip()
                    if code != 'NO_CODE' and self.unspsc.exists(code):
                        claude_top4_codes.add(code)
            
            logger.debug(f"Top 4 codes - GPT: {gpt_top4_codes}, Gemini: {gemini_top4_codes}, Claude: {claude_top4_codes}")
            
            # STEP 4: Find common codes and build shortlist
            result.common_codes = self._find_common_codes(gpt_codes, gemini_codes, claude_codes)
            logger.debug(f"Found {len(result.common_codes)} common codes (3 out of 3)")
            
            result.shortlist_for_rerank = self._build_shortlist(
                result.common_codes,
                result.gpt_candidates
            )
            logger.debug(f"Shortlist for reranking: {result.shortlist_for_rerank}")
            
            # Build Top 5 Recommendations from ALL candidates
            # NEW REQUIREMENT: Uses all candidates from all 3 LLMs (including Top 4)
            all_candidates = result.gpt_candidates + result.gemini_candidates + result.claude_candidates
            logger.debug(f"Building Top 5 from ALL {len(all_candidates)} candidates by confidence")
            result.top_5_recommendations = self._build_top_5_recommendations(
            all_candidates,
            result.gpt_top4_with_codes,
            result.gemini_top4_with_codes,
            result.claude_top4_with_codes,
            )
            
            # Rank all 30 candidates by confidence (excluding Top 5)
            logger.debug(f"Ranking all candidates by confidence (excluding Top 5)")
            result.all_30_ranked = self._rank_all_30_by_confidence(
                all_candidates,
                result.top_5_recommendations  # Pass Top 5 to exclude them
            )
            
            # Check if we have a shortlist
            if not result.shortlist_for_rerank:
                logger.warning(f"No shortlist for {material_code} - using fallback from Top 5")
                result.decision_rule = "no_shortlist_fallback"
                
                # FALLBACK: Use the first code from Top 5 Recommendations
                if result.top_5_recommendations:
                    result.final_recommendation = result.top_5_recommendations[0]
                    hierarchy = self.unspsc.get_hierarchy(result.final_recommendation)
                    result.final_commodity = hierarchy.get('title')
                    result.segment = hierarchy.get('segment')
                    result.family = hierarchy.get('family')
                    result.class_name = hierarchy.get('class')
                    logger.info(f"{material_code}: Using Top 5 fallback -> {result.final_recommendation}")
                else:
                    result.decision_rule = "no_shortlist_no_fallback"
                    logger.warning(f"{material_code}: No shortlist and no Top 5 recommendations available")
                
                result.processing_time = time.time() - start_time
                return result
            
            # TIER 2: Reranking (pick best from shortlist)
            logger.debug(f"Tier 2: Reranking for {material_code}")
            
            tier2_system = PromptTemplates.TIER2_SYSTEM
            tier2_user = PromptTemplates.tier2_user_prompt(
                material_description,
                po_text,
                result.shortlist_for_rerank,
                self.unspsc,
                item_detail
            )
            
            tier2_responses = self._call_providers_tier2(tier2_system, tier2_user)
            
            # Extract final picks
            gpt_resp = tier2_responses.get('chatgpt', ProviderResponse(provider='chatgpt'))
            gemini_resp = tier2_responses.get('gemini', ProviderResponse(provider='gemini'))
            claude_resp = tier2_responses.get('claude', ProviderResponse(provider='claude'))
            
            result.final_gpt_code = self._normalize_code(gpt_resp.candidates[0].code) if gpt_resp.candidates else None
            result.final_gemini_code = self._normalize_code(gemini_resp.candidates[0].code) if gemini_resp.candidates else None
            result.final_claude_code = self._normalize_code(claude_resp.candidates[0].code) if claude_resp.candidates else None
            
            # Log any errors from Tier 2
            if gpt_resp.error:
                logger.warning(f"{material_code}: ChatGPT Tier2 error: {gpt_resp.error}")
            if gemini_resp.error:
                logger.warning(f"{material_code}: Gemini Tier2 error: {gemini_resp.error}")
            if claude_resp.error:
                logger.warning(f"{material_code}: Claude Tier2 error: {claude_resp.error}")
            
            # Debug log the codes before consensus
            logger.debug(f"{material_code} Tier2 codes: GPT={result.final_gpt_code}, Gemini={result.final_gemini_code}, Claude={result.final_claude_code}")
            
            # STEP 5: Apply consensus
            result.final_recommendation, result.decision_rule = self._apply_consensus(
                result.final_gpt_code,
                result.final_gemini_code,
                result.final_claude_code
            )
            
            # Debug log the consensus result
            logger.debug(f"{material_code} Consensus: recommendation={result.final_recommendation}, rule={result.decision_rule}")
            
            # Get hierarchy and title
            if result.final_recommendation:
                hierarchy = self.unspsc.get_hierarchy(result.final_recommendation)
                result.final_commodity = hierarchy.get('title')
                result.segment = hierarchy.get('segment')
                result.family = hierarchy.get('family')
                result.class_name = hierarchy.get('class')
                
                # Calculate confidence (average of final picks)
                confidences = []
                if gpt_resp.candidates:
                    confidences.append(gpt_resp.candidates[0].confidence)
                if gemini_resp.candidates:
                    confidences.append(gemini_resp.candidates[0].confidence)
                if claude_resp.candidates:
                    confidences.append(claude_resp.candidates[0].confidence)
                
                result.confidence = sum(confidences) / len(confidences) if confidences else 0.0
            else:
                # FALLBACK: If consensus returned None, use Top 5 Recommendations
                logger.warning(f"{material_code}: Consensus returned None - using Top 5 fallback")
                if result.top_5_recommendations:
                    result.final_recommendation = result.top_5_recommendations[0]
                    result.decision_rule = f"{result.decision_rule}_top5_fallback"
                    hierarchy = self.unspsc.get_hierarchy(result.final_recommendation)
                    result.final_commodity = hierarchy.get('title')
                    result.segment = hierarchy.get('segment')
                    result.family = hierarchy.get('family')
                    result.class_name = hierarchy.get('class')
                    result.confidence = 0.5  # Medium confidence for fallback
                    logger.info(f"{material_code}: Using Top 5 fallback -> {result.final_recommendation}")
                else:
                    logger.error(f"{material_code}: No consensus and no Top 5 recommendations available")
            
            result.processing_time = time.time() - start_time
            
            # Cache result
            self.cache.set(cache_key, result.__dict__)
            
            return result
            
        except Exception as e:
            logger.error(f"Error classifying {material_code}: {e}")
            result.decision_rule = f"error: {str(e)}"
            result.processing_time = time.time() - start_time
            return result
    
    def classify_batch(
        self,
        df: pd.DataFrame,
        material_col: str = "Material",
        description_col: str = "Material_Description",
        po_col: str = "PO_Text",
        challenge_col: str = "Challenge_Level",
        item_detail_col: str = "Item_detail",
        matl_group_col: str = "Matl Group",
        plant_col: str = "Plant",
        output_file: str = "output.xlsx"
    ) -> pd.DataFrame:
        """Classify a batch of materials."""
        
        logger.info(f"Starting batch classification of {len(df)} materials")
        
        results = []
        
        # Progress bar
        with tqdm(total=len(df), desc="Classifying materials") as pbar:
            for idx, row in df.iterrows():
                material_code = str(row.get(material_col, f"MAT_{idx}"))
                material_desc = str(row.get(description_col, ""))
                po_text = str(row.get(po_col, ""))
                challenge = str(row.get(challenge_col, "Medium"))
                item_detail = str(row.get(item_detail_col, ""))
                matl_group = str(row.get(matl_group_col, ""))
                plant = str(row.get(plant_col, ""))
                
                result = self.classify_material(
                    material_code,
                    material_desc,
                    po_text,
                    challenge,
                    item_detail,
                    matl_group,
                    plant
                )
                
                results.append(result)
                pbar.update(1)
                
                # Log progress every 100 materials
                if (idx + 1) % 100 == 0:
                    logger.info(f"Processed {idx + 1}/{len(df)} materials")
        
        # Convert to DataFrame
        output_df = self._create_output_dataframe(results)
        
        # Save to Excel with formatting
        logger.info(f"Saving results to {output_file}")
        
        # Use ExcelWriter for better formatting control
        try:
            from openpyxl import load_workbook
            from openpyxl.styles import Alignment
            
            # Save dataframe first
            output_df.to_excel(output_file, index=False, engine='openpyxl')
            
            # Open and format
            wb = load_workbook(output_file)
            ws = wb.active
            
            # Columns that need text wrapping (multi-line content)
            wrap_columns = ['GPT_Commodity', 'Gemini_Commodity', 'Claude_Commodity',
                          'GPT_Top4_With_Codes', 'Gemini_Top4_With_Codes', 'Claude_Top4_With_Codes',
                          'All_Commodities_Ranked',
                          'Common_Commodity', 'Top_5_Recommendations', 'All_30_Ranked_By_Confidence']
            
            # Get column indices for wrap columns
            header_row = [cell.value for cell in ws[1]]
            wrap_col_indices = [header_row.index(col) + 1 for col in wrap_columns if col in header_row]
            
            # Apply text wrapping and alignment
            for row_idx, row in enumerate(ws.iter_rows(min_row=2), start=2):  # Skip header
                for col_idx, cell in enumerate(row, start=1):
                    if col_idx in wrap_col_indices:
                        # Enable text wrapping
                        cell.alignment = Alignment(
                            wrap_text=True,
                            vertical='top',
                            horizontal='left'
                        )
            
            # Auto-adjust column widths
            for column in ws.columns:
                max_length = 0
                column_letter = column[0].column_letter
                
                # Get column index
                col_idx = column[0].column
                
                # Set specific widths for different columns
                if col_idx in wrap_col_indices:
                    ws.column_dimensions[column_letter].width = 50  # Wide for multi-line
                elif header_row[col_idx - 1] in ['Material_Description', 'PO_Text']:
                    ws.column_dimensions[column_letter].width = 40
                elif header_row[col_idx - 1] in ['Final_Recommendation']:
                    ws.column_dimensions[column_letter].width = 40
                else:
                    ws.column_dimensions[column_letter].width = 15
            
            # Save formatted workbook
            wb.save(output_file)
            logger.info(f"Applied Excel formatting (text wrapping, column widths)")
            
        except ImportError:
            # Fallback if openpyxl formatting not available
            logger.warning("openpyxl not available for advanced formatting. Saving basic Excel file.")
            output_df.to_excel(output_file, index=False)
        except Exception as e:
            logger.error(f"Error applying Excel formatting: {e}. Saved basic file.")
        
        logger.info(f"Batch classification complete!")
        
        return output_df
    
    def _create_output_dataframe(self, results: List[ClassificationResult]) -> pd.DataFrame:
        """Create output DataFrame matching desired format."""
        
        def format_candidates(candidates: List[Candidate]) -> str:
            """Format candidates as 'code - commodity' string with newlines.
            
            IMPORTANT: Only includes candidates with VALID UNSPSC codes.
            """
            formatted = []
            for c in candidates[:10]:
                code = self._normalize_code(c.code)
                if code and self.unspsc.exists(code):  # VALIDATE CODE
                    title = self.unspsc.get_title(code) or c.commodity
                    formatted.append(f"{code} - {title}")
            return "\n".join(formatted) if formatted else ""
        
        def format_code_list(codes: List[str]) -> str:
            """Format code list with titles using newlines."""
            formatted = []
            for code in codes[:5]:
                title = self.unspsc.get_title(code) or "(unknown)"
                formatted.append(f"{code} - {title}")
            return "\n".join(formatted) if formatted else ""
        
        def format_single_code(code: Optional[str]) -> str:
            """Format single code with title."""
            if not code:
                return ""
            title = self.unspsc.get_title(code) or "(unknown)"
            return f"{code} - {title}"
        
        rows = []
        for r in results:
            rows.append({
                'Date': r.timestamp,
                'Material_Code': r.material_code,
                'Material_Description': r.material_description,
                'PO_Text': r.po_text,
                'Item_Detail': r.item_detail,
                'Matl_Group': r.matl_group,
                'Plant': r.plant,
                'GPT_Commodity': format_candidates(r.gpt_candidates),
                'Gemini_Commodity': format_candidates(r.gemini_candidates),
                'Claude_Commodity': format_candidates(r.claude_candidates),
                # NEW V8: Top 4 commodities with codes from each LLM
                'GPT_Top4_With_Codes': "\n".join(r.gpt_top4_with_codes) if r.gpt_top4_with_codes else "",
                'Gemini_Top4_With_Codes': "\n".join(r.gemini_top4_with_codes) if r.gemini_top4_with_codes else "",
                'Claude_Top4_With_Codes': "\n".join(r.claude_top4_with_codes) if r.claude_top4_with_codes else "",
                # NEW V8: All commodities ranked by confidence
                'All_Commodities_Ranked': "\n".join(r.all_commodities_ranked) if r.all_commodities_ranked else "",
                'Common_Commodity': format_code_list(r.common_codes),
                'Top_5_Recommendations': format_code_list(r.top_5_recommendations),
                'All_30_Ranked_By_Confidence': "\n".join(r.all_30_ranked) if r.all_30_ranked else "",  # NEW COLUMN
                'Final_GPT_Commodity': format_single_code(r.final_gpt_code),
                'Final_Gemini_Commodity': format_single_code(r.final_gemini_code),
                'Final_Claude_Commodity': format_single_code(r.final_claude_code),
                'Final_Recommendation': format_single_code(r.final_recommendation),
                'Commodity_Code': r.final_recommendation or "",
                'Segment': r.segment or "",
                'Family': r.family or "",
                'Class': r.class_name or "",
                'Decision_Rule': r.decision_rule,
                'Confidence': r.confidence,
                'Processing_Time': round(r.processing_time, 2),
                'From_Cache': r.from_cache
            })
        
        return pd.DataFrame(rows)


# =============================================================================
# Main Entry Point
# =============================================================================

def main():
    """Main entry point for the pipeline."""
    
    parser = argparse.ArgumentParser(
        description="UNSPSC Classification Pipeline - Production v1.0",
        formatter_class=argparse.RawDescriptionHelpFormatter,
        epilog="""
Examples:
  # Basic usage (with command-line arguments)
  python %(prog)s --input materials.xlsx --output results.xlsx --unspsc codes.csv
  
  # With caching
  python %(prog)s --input materials.xlsx --output results.xlsx --unspsc codes.csv --cache cache.jsonl
  
  # Using default paths (configured in DEFAULT_PATHS at top of script)
  python %(prog)s
  
  # Override only specific paths
  python %(prog)s --input my_materials.xlsx
  
  # Sequential processing (no parallel)
  python %(prog)s --input materials.xlsx --output results.xlsx --unspsc codes.csv --no-parallel
  
  # Verbose logging
  python %(prog)s --input materials.xlsx --output results.xlsx --unspsc codes.csv --verbose

Environment Variables:
  OPENAI_API_KEY      OpenAI API key for ChatGPT
  ANTHROPIC_API_KEY   Anthropic API key for Claude
  GOOGLE_API_KEY      Google API key for Gemini

Note: 
  - All 3 API keys must be set in environment variables
  - Configure DEFAULT_PATHS at top of script for easy repeated runs
  - Command-line arguments override DEFAULT_PATHS
        """
    )
    
    # Required arguments (can use defaults if configured)
    parser.add_argument(
        '--input', 
        default=DEFAULT_PATHS.get("input"),
        help=f'Input Excel/CSV file with materials (default: {DEFAULT_PATHS.get("input", "required")})'
    )
    parser.add_argument(
        '--output', 
        default=DEFAULT_PATHS.get("output"),
        help=f'Output Excel file for results (default: {DEFAULT_PATHS.get("output", "required")})'
    )
    parser.add_argument(
        '--unspsc', 
        default=DEFAULT_PATHS.get("unspsc"),
        help=f'UNSPSC codes CSV file (default: {DEFAULT_PATHS.get("unspsc", "required")})'
    )
    
    # Optional arguments
    parser.add_argument(
        '--cache', 
        default=DEFAULT_PATHS.get("cache"),
        help=f'JSONL cache file (default: {DEFAULT_PATHS.get("cache", "none")})'
    )
    parser.add_argument('--no-parallel', action='store_true', help='Disable parallel API calls')
    parser.add_argument('--verbose', action='store_true', help='Enable verbose logging')
    
    # Column names (with defaults)
    parser.add_argument('--material-col', default=DEFAULT_CONFIG.get("material_col", "Material"), help='Material code column (default: Material)')
    parser.add_argument('--description-col', default=DEFAULT_CONFIG.get("description_col", "Material_Description"), help='Description column (default: Material_Description)')
    parser.add_argument('--po-col', default=DEFAULT_CONFIG.get("po_col", "PO_Text"), help='PO text column (default: PO_Text)')
    parser.add_argument('--challenge-col', default=DEFAULT_CONFIG.get("challenge_col", "Challenge_Level"), help='Challenge level column (default: Challenge_Level)')
    parser.add_argument('--item-detail-col', default=DEFAULT_CONFIG.get("item_detail_col", "Item_detail"), help='Item detail column (default: Item_detail)')
    parser.add_argument('--matl-group-col', default=DEFAULT_CONFIG.get("matl_group_col", "Matl Group"), help='Material Group column (default: Matl Group)')
    parser.add_argument('--plant-col', default=DEFAULT_CONFIG.get("plant_col", "Plant"), help='Plant column (default: Plant)')
    
    # Model selection
    parser.add_argument('--gpt-model', default=DEFAULT_MODELS.get("gpt", "gpt-4.1-2025-04-14"), help='ChatGPT model (default: gpt-4.1-2025-04-14)')
    parser.add_argument('--gemini-model', default=DEFAULT_MODELS.get("gemini", "gemini-2.0-flash-001"), help='Gemini model (default: gemini-2.0-flash-001)')
    parser.add_argument('--claude-model', default=DEFAULT_MODELS.get("claude", "claude-haiku-4-5-20251001"), help='Claude model (default: claude-haiku-4-5-20251001)')
    
    args = parser.parse_args()
    
    # Setup logging
    setup_logging(args.verbose)
    
    logger.info("=" * 80)
    logger.info("UNSPSC Classification Pipeline - Production v1.0")
    logger.info("=" * 80)
    
    # Check API keys
    openai_key = os.getenv('OPENAI_API_KEY')
    anthropic_key = os.getenv('ANTHROPIC_API_KEY')
    google_key = os.getenv('GOOGLE_API_KEY')
    
    if not all([openai_key, anthropic_key, google_key]):
        logger.error("ERROR: Missing API keys!")
        logger.error("Please set environment variables:")
        if not openai_key:
            logger.error("  - OPENAI_API_KEY")
        if not anthropic_key:
            logger.error("  - ANTHROPIC_API_KEY")
        if not google_key:
            logger.error("  - GOOGLE_API_KEY")
        sys.exit(1)
    
    # logger.info("✓ All API keys found")
    # New, safe version
    logger.info("[OK] All API keys found")

    # Validate required paths
    if not args.input:
        logger.error("ERROR: --input path is required!")
        logger.error("Either provide --input argument or set DEFAULT_PATHS['input'] in the script")
        sys.exit(1)
    
    if not args.output:
        logger.error("ERROR: --output path is required!")
        logger.error("Either provide --output argument or set DEFAULT_PATHS['output'] in the script")
        sys.exit(1)
    
    if not args.unspsc:
        logger.error("ERROR: --unspsc path is required!")
        logger.error("Either provide --unspsc argument or set DEFAULT_PATHS['unspsc'] in the script")
        sys.exit(1)
    
    logger.info(f"Configuration:")
    logger.info(f"  Input: {args.input}")
    logger.info(f"  Output: {args.output}")
    logger.info(f"  UNSPSC: {args.unspsc}")
    if args.cache:
        logger.info(f"  Cache: {args.cache}")
    
    # Load UNSPSC dictionary
    logger.info(f"Loading UNSPSC codes from {args.unspsc}")
    unspsc_dict = UNSPSCDict(args.unspsc)
    
    # Initialize cache
    cache = JSONLCache(args.cache)
    
    # Initialize providers
    chatgpt = ChatGPTProvider(openai_key, args.gpt_model)
    gemini = GeminiProvider(google_key, args.gemini_model)
    claude = ClaudeProvider(anthropic_key, args.claude_model)
    
    logger.info(f"Models: GPT={args.gpt_model}, Gemini={args.gemini_model}, Claude={args.claude_model}")
    
    # Initialize pipeline
    pipeline = UNSPSCPipeline(
        unspsc_dict=unspsc_dict,
        cache=cache,
        chatgpt=chatgpt,
        gemini=gemini,
        claude=claude,
        parallel=not args.no_parallel
    )
    
    # Load input data
    logger.info(f"Loading input data from {args.input}")
    input_path = Path(args.input)
    
    if not input_path.exists():
        logger.error(f"ERROR: Input file not found: {args.input}")
        sys.exit(1)
    
    try:
        if input_path.suffix.lower() in ['.xlsx', '.xls']:
            df = pd.read_excel(args.input)
        else:
            df = pd.read_csv(args.input)
        
        logger.info(f"Loaded {len(df)} materials")
        
    except Exception as e:
        logger.error(f"ERROR loading input file: {e}")
        sys.exit(1)
    
    # Run classification
    try:
        output_df = pipeline.classify_batch(
            df,
            material_col=args.material_col,
            description_col=args.description_col,
            po_col=args.po_col,
            challenge_col=args.challenge_col,
            item_detail_col=args.item_detail_col,
            matl_group_col=args.matl_group_col,
            plant_col=args.plant_col,
            output_file=args.output
        )
        
        logger.info("=" * 80)
        logger.info("CLASSIFICATION COMPLETE!")
        logger.info(f"Results saved to: {args.output}")
        logger.info(f"Total materials processed: {len(output_df)}")
        logger.info(f"From cache: {output_df['From_Cache'].sum()}")
        logger.info(f"Newly processed: {(~output_df['From_Cache']).sum()}")
        logger.info("=" * 80)
        
    except Exception as e:
        logger.error(f"ERROR during classification: {e}")
        logger.exception("Full traceback:")
        sys.exit(1)


if __name__ == "__main__":
    main()